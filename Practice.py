
# https://www.plus2net.com/python/tkinter-treeview-excel.php


from openpyxl import load_workbook
from tkinter import ttk
from tkinter import *


# 엑셀 데이터 추출
wb = load_workbook("C:/Python/Code/ToDoList/student.xlsx")
ws = wb['student']
# values_only=True를 해야 위치값이 아닌 실제 데이터 값이 도출됨
header = ws.iter_rows(min_row=1, max_row=1, max_col=5, values_only=True)
datas = ws.iter_rows(min_row=2, max_col=5, values_only=True)
# print(header)
header = [r for r in header]
datas = [r for r in datas]
wb.close()
# print(header[0])
# datas[0]


# tkinter 활용
root = Tk()
root.title("Tkinter Pratice loading excel file")
# root.geometry("560x280") # width and hight of window 

# treeview 활용
# y,x 스크롤바 추가 필요
tree = ttk.Treeview(root, selectmode='browse')
tree.pack(expand=True, fill="both")
# tree.grid(row=0,column=0,columnspan=3,padx=30,pady=20)

# Number of rows to display, default is 10
tree['height'] = 20
# 컬럼 제목만 보이게함?
tree['show'] = 'headings'

tree['columns'] = header[0]

for i in header[0]:
    tree.column(i, width=100, anchor='c')
    tree.heading(i, text=i, anchor='c')


for data in datas:
    tree.insert("",'end', iid=data[0], values=data)


root.mainloop()















