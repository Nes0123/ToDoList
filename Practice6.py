from openpyxl import load_workbook

wb = load_workbook("C:/Python/Code/ToDoList/ToDoList_Form.xlsx")
db_ws = wb['DB']

def wordfinder(SearchString):
    for i in range(1, db_ws.max_row+1):
        if SearchString == db_ws.cell(i,1).value:
            print(i)
    return i

wordfinder()