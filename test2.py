from openpyxl import load_workbook
from tkinter import ttk
import tkinter.ttk as ttk
from tkinter import *
import tkinter as tk 
import win32com.client
import pandas as pd
import tkinter.messagebox as msgbox
import os
import tkinter.font as font
import time



def sort():
    df_db = pd.read_excel("C:/Python/Code/ToDoList/ToDoList_Form.xlsx", sheet_name='DB')
    # print(db_df)
    df_db = df_db.sort_values(by=['No'])

    # 넘버링 새롭게 하기
    lv1_cnt = 0
    for i in range(len(df_db['업무레벨'])):
        # print(df_db['업무레벨'][i])
        if df_db['업무레벨'][i] == 1 :
            lv1_cnt += 1
    print(lv1_cnt)
    print(int(df_db['No'][len(df_db['No'])-1][0:2]))
    print(type(lv1_cnt))
    if lv1_cnt == 7:
        print("123")
    if int(df_db['No'][len(df_db['No'])-1][0:2]) == 7:
        print("445")
    lv1_cnt = 0
    if lv1_cnt != int(df_db['No'][len(df_db['No'])-1][0:2]) :
        print("different")
        for i in range(len(df_db['업무레벨'])):
        # print(df_db['업무레벨'][i])
            if df_db['업무레벨'][i] == 1 :
                lv1_cnt += 1
            # print("asis : ", df_db['No'][i])
            df_db['No'][i] = format(lv1_cnt,'02') + "-"+ df_db['No'][i][3:]
            # print("tobe : ", df_db['No'][i])
            
            # df_db['No'][len(df_db['No'])-1][0:2] = format(lv1_cnt,'02')
                
    # print("lv1_cnt : ", lv1_cnt)
    # print(db_df)
    df_db.to_excel("C:/Python/Code/ToDoList/test.xlsx",index=False, sheet_name="DB")
    path1 = "C:/Python/Code/ToDoList/ToDoList_Form.xlsx"
    path2 = "C:/Python/Code/ToDoList/test.xlsx"
    
    wb2 = load_workbook(filename=path2)
    ws2 = wb2['DB']

    wb1 = load_workbook(filename=path1)
    wb1.remove(wb1['DB'])
    ws1 = wb1.create_sheet()
    ws1.title = 'DB'
    # ws1 = wb1['DB']
    for row in ws2:
        for cell in row:
            ws1[cell.coordinate].value = cell.value
    wb1.save(path1)


sort()