
from openpyxl import load_workbook
from tkinter import ttk
from tkinter import *

# Database load
def load_db():

    # 엑셀 데이터 추출
    wb = load_workbook("C:/Python/Code/ToDoList/ToDoList_Form.xlsx")
    db_ws = wb['DB']
    # values_only=True를 해야 위치값이 아닌 실제 데이터 값이 도출됨
    db_header = db_ws.iter_rows(min_row=1, max_row=1, max_col=9, values_only=True)
    db_datas = db_ws.iter_rows(min_row=2, max_col=9, values_only=True)

    db_header = [r for r in db_header]
    db_datas = [r for r in db_datas]
    wb.close()
    return(wb, db_ws, db_header, db_datas) # ??되나?
 
def add_task():
    # 업무 레벨에 따라 No링 잘 하기
    # no 정렬 방식 활용 해보기
    add_tk = Tk()
    add_tk.title("업무 추가")

    def add_db():
        wb = load_workbook("C:/Python/Code/ToDoList/ToDoList_Form.xlsx")
        db_ws = wb['DB']
        # values_only=True를 해야 위치값이 아닌 실제 데이터 값이 도출됨
        db_header = db_ws.iter_rows(min_row=1, max_row=1, max_col=9, values_only=True)
        db_datas = db_ws.iter_rows(min_row=2, max_col=9, values_only=True)
        db_header = [r for r in db_header]
        db_datas = [r for r in db_datas]
        
        # print(db_datas[db_ws.max_row-2][0])
        lv1_cnt = int(db_datas[db_ws.max_row-2][0][0:2])
        lv1_next = lv1_cnt+1
        lv1_next = format(lv1_next,'02')
        # print(format(lv1_next,'02'))
        # print(db_ws.max_row)
        r_add = db_ws.max_row + 1
        db_ws.cell(row=r_add, column=1).value = lv1_next + "-00-00"
        db_ws.cell(row=r_add, column=2).value = entry_task_name.get()
        db_ws.cell(row=r_add, column=3).value = entry_team_name.get()
        db_ws.cell(row=r_add, column=4).value = entry_person_name.get()
        db_ws.cell(row=r_add, column=5).value = cmb_situation.get()
        db_ws.cell(row=r_add, column=6).value = entry_date.get()
        db_ws.cell(row=r_add, column=7).value = entry_time.get()
        db_ws.cell(row=r_add, column=8).value = entry_note.get()
        db_ws.cell(row=r_add, column=9).value = 1

        wb.save("C:/Python/Code/ToDoList/ToDoList_Form.xlsx")


    # 업무명/내용
    frame_task_name = Frame(add_tk)
    frame_task_name.pack(fill="x")

    lbl_task_name = Label(frame_task_name, text="업무명/내용 - 레벨1")
    lbl_task_name.pack(side="left")

    entry_task_name = Entry(frame_task_name)
    entry_task_name.pack(side="right")

    # 담당팀
    frame_team_name = Frame(add_tk)
    frame_team_name.pack(fill="x")

    lbl_team_name = Label(frame_team_name, text="담당팀")
    lbl_team_name.pack(side="left")

    entry_team_name = Entry(frame_team_name)
    entry_team_name.pack(side="right")

    # 담당자
    frame_person_name = Frame(add_tk)
    frame_person_name.pack(fill="x")

    lbl_person_name = Label(frame_person_name, text="담당자")
    lbl_person_name.pack(side="left")

    entry_person_name = Entry(frame_person_name)
    entry_person_name.pack(side="right")

    # 상황
    frame_situation = Frame(add_tk)
    frame_situation.pack(fill="x")

    lbl_situation = Label(frame_situation, text="상황")
    lbl_situation.pack(side="left")

    opt_situation = ["진행","완료","검토","취소","중단","지연"]
    cmb_situation = ttk.Combobox(frame_situation,state="readonly",
    values=opt_situation)
    cmb_situation.current(0)
    cmb_situation.pack(side="right")

    # 완료날짜
    frame_date = Frame(add_tk)
    frame_date.pack(fill="x")

    lbl_date = Label(frame_date, text="완료날짜")
    lbl_date.pack(side="left")

    entry_date = Entry(frame_date)
    entry_date.pack(side="right")
    entry_date.insert(END, "2022-12-01")

    # 완료시간
    frame_time = Frame(add_tk)
    frame_time.pack(fill="x")

    lbl_time = Label(frame_time, text="완료시간")
    lbl_time.pack(side="left")

    entry_time = Entry(frame_time)
    entry_time.pack(side="right")
    entry_time.insert(END, "13:00")

    
    # 비고
    frame_note = Frame(add_tk)
    frame_note.pack(fill="x")

    lbl_note = Label(frame_note, text="비고")
    lbl_note.pack(side="left")

    entry_note = Entry(frame_note)
    entry_note.pack(side="right")

    # fucntion2 frame
    frame_func2 = Frame(add_tk)
    frame_func2.pack(fill="x")

    # add db
    btn_add_db = Button(frame_func2, text="추가",command=add_db)
    btn_add_db.pack(side="left")

    # cancel button
    # quit 함수를 쓰면 전체 프로그램이 종료되어서 withdraw 함수 사용함
    btn_cancel = Button(frame_func2, text="취소", command=add_tk.withdraw)
    btn_cancel.pack(side="right")

    add_tk.mainloop()


# 첫 실행 시 Database load

wb = load_workbook("C:/Python/Code/ToDoList/ToDoList_Form.xlsx")
db_ws = wb['DB']
# values_only=True를 해야 위치값이 아닌 실제 데이터 값이 도출됨
db_header = db_ws.iter_rows(min_row=1, max_row=1, max_col=9, values_only=True)
db_datas = db_ws.iter_rows(min_row=2, max_col=9, values_only=True)

db_header = [r for r in db_header]
db_datas = [r for r in db_datas]
wb.close()

# 프로그램 메인 화면 tkinter 

win = Tk()
win.title("To Do List")

# treeview 활용
# y,x 스크롤바 추가 필요
# 클릭시 여러개 선택 가능하도록 (extended)
tree = ttk.Treeview(win, selectmode='extended')
# tree = ttk.Treeview(win, selectmode='browse')

tree.pack(expand=True, fill="both")

# # Number of rows to display, default is 10
tree['height'] = 20

tree['columns'] = ("1", "2", "3", "4", "5", "6")

# column보다 heading이 1개 더 많아서(#0) 하이라키구조 처럼 +- 표시가 됨. 
tree.heading("#0", text="업무명/내용")
tree.heading("#1", text="담당팀")
tree.heading("#2", text="담당자")
tree.heading("#3", text="상황")
tree.heading("#4", text="완료날짜")
tree.heading("#5", text="완료시간")
tree.heading("#6", text="비고")

tree.column("1", width=100)
tree.column("2", width=100)
tree.column("3", width=50, anchor='c')
tree.column("4", width=70, anchor='c')
tree.column("5", width=70, anchor='c')
tree.column("6", width=100)


for data in db_datas:
    # print(data[0][6:8])
    # text 값이 있어야 하이라키구조 처럼 보여질 수 있음 (+-표시).
    # lv1에만 open=true를 줘서 실행 시 lv2까지만 자동 보여주기    
    if data[0][3:8] == '00-00':
        lv1 = tree.insert('','end', iid=data[0], text=data[1], values=data[2:8], open=True)
    elif '00' in data[0][6:8]:
        lv2 = tree.insert(lv1,'end',iid=data[0], text=data[1] ,values=data[2:8])
    else :
        lv3 = tree.insert(lv2,'end',iid=data[0], text=data[1] ,values=data[2:8])


# function frame
func_frame = Frame(win)
func_frame.pack(fill="x")

# 업무 추가 버튼
btn_add_task = Button(func_frame, text="업무추가", command=add_task)
btn_add_task.pack(side="left")

# 종료 버튼
btn_close = Button(func_frame, text="종료", command=win.quit)
btn_close.pack(side="right")

win.mainloop()

