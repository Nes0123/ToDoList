
from openpyxl import load_workbook
from tkinter import ttk
import tkinter.ttk as ttk
from tkinter import *
import tkinter as tk 
import win32com.client
import pandas as pd
import tkinter.messagebox as msgbox
import os
import tkinter.font as font
import time


# 필요 : class 사용하기, file name(os.getcwd)
# def sort(): 의 과제 넘버링 다시 해서 빈 번호 채우기.
# 담당팀, 담당자 주소록 만들기, 편집, 삭제 기능
# x스크롤바 추가
# 즐겨찾기 편집(담당자)의 삭제? 이런 것이 문제있었던 것 같음
# 하위업무의 담당자나 담당팀 추가 프레임이 직관적이지 않음 

def font_define():
    func_font = font.Font(family='맑은 고딕', size=12)
    return func_font

def add_task():
    try:
        # 업무 레벨에 따라 No링 잘 하기
        # no 정렬 방식 활용 해보기
        add_tk = Tk()
        add_tk.title("업무 추가")
        add_tk.geometry("350x230")
        def add_db():
            wb = load_workbook("C:/Python/Code/ToDoList/ToDoList_Form.xlsx")
            db_ws = wb['DB']
            # values_only=True를 해야 위치값이 아닌 실제 데이터 값이 도출됨
            db_header = db_ws.iter_rows(min_row=1, max_row=1, max_col=9, values_only=True)
            db_datas = db_ws.iter_rows(min_row=2, max_col=9, values_only=True)
            db_header = [r for r in db_header]
            db_datas = [r for r in db_datas]
            
            # print(db_datas[db_ws.max_row-2][0])
            lv1_cnt = int(db_datas[db_ws.max_row-2][0][0:2])
            # print(lv1_cnt)
            lv1_next = lv1_cnt+1
            lv1_next = format(lv1_next,'02')
            # print(format(lv1_next,'02'))
            # print(db_ws.max_row)
            r_add = db_ws.max_row + 1
            db_ws.cell(row=r_add, column=1).value = lv1_next + "-00-00"
            # 값이 없을 때 공백 하나 넣기. 값이 없으면 편집 함수 중 entry 값을 불러올 때 에러가 남
            if len(entry_task_name.get()) == 0 :
                db_ws.cell(row=r_add, column=2).value = " "    
            else : 
                db_ws.cell(row=r_add, column=2).value = entry_task_name.get()
            if len(entry_team_name.get()) == 0 :
                db_ws.cell(row=r_add, column=3).value = " "
            else :    
                db_ws.cell(row=r_add, column=3).value = entry_team_name.get()
            if len(entry_person_name.get()) == 0:
                db_ws.cell(row=r_add, column=4).value = " "
            else :    
                db_ws.cell(row=r_add, column=4).value = entry_person_name.get()
            db_ws.cell(row=r_add, column=5).value = cmb_situation.get()
            if len(entry_date.get()) == 0:
                db_ws.cell(row=r_add, column=6).value = " "    
            else :
                db_ws.cell(row=r_add, column=6).value = entry_date.get()
            if len(entry_time.get()) == 0:
                db_ws.cell(row=r_add, column=7).value = " "
            else :    
                db_ws.cell(row=r_add, column=7).value = entry_time.get()
            if len(entry_note.get()) == 0:
                db_ws.cell(row=r_add, column=8).value = " "
            else :     
                db_ws.cell(row=r_add, column=8).value = entry_note.get()
            db_ws.cell(row=r_add, column=9).value = 1

            wb.save("C:/Python/Code/ToDoList/ToDoList_Form.xlsx")

            msgbox.showinfo("알림", "신규 업무 추가가 완료되었습니다.")

            add_tk.withdraw()
            sort()
            load_task()

        # func_font3 = font.Font(family='궁서', size=8)


        # 업무명/내용
        frame_task_name = Frame(add_tk)
        frame_task_name.pack(fill="x")

        lbl_task_name = Label(frame_task_name, text="업무명/내용 - 레벨1", font=('맑은 고딕',12))
        # lbl_task_name = Label(frame_task_name, text="업무명/내용 - 레벨1", font=('맑은 고딕',12,'bold'))
        # lbl_task_name['font'] = func_font2
        lbl_task_name.pack(side="left")

        entry_task_name = Entry(frame_task_name, font=('맑은 고딕',12))
        entry_task_name.pack(side="right")

        # 담당팀
        frame_team_name = Frame(add_tk)
        frame_team_name.pack(fill="x")

        lbl_team_name = Label(frame_team_name, text="담당팀", font=('맑은 고딕',12))
        # lbl_team_name['font'] = func_font3
        lbl_team_name.pack(side="left")

        entry_team_name = Entry(frame_team_name, font=('맑은 고딕',12))
        entry_team_name.pack(side="right")

        # 담당자
        frame_person_name = Frame(add_tk)
        frame_person_name.pack(fill="x")

        lbl_person_name = Label(frame_person_name, text="담당자", font=('맑은 고딕',12))
        lbl_person_name.pack(side="left")

        entry_person_name = Entry(frame_person_name, font=('맑은 고딕',12))
        entry_person_name.pack(side="right")

        # 상황
        frame_situation = Frame(add_tk)
        frame_situation.pack(fill="x")

        lbl_situation = Label(frame_situation, text="상황", font=('맑은 고딕',12))
        lbl_situation.pack(side="left")

        opt_situation = ["진행","완료","검토","취소","중단","지연"]
        cmb_situation = ttk.Combobox(frame_situation,state="readonly",
        values=opt_situation, font=('맑은 고딕',12))
        cmb_situation.current(0)
        cmb_situation.pack(side="right")

        # 완료날짜
        frame_date = Frame(add_tk)
        frame_date.pack(fill="x")

        lbl_date = Label(frame_date, text="완료날짜", font=('맑은 고딕',12))
        lbl_date.pack(side="left")

        entry_date = Entry(frame_date, font=('맑은 고딕',12))
        entry_date.pack(side="right")
        # entry_date.insert(END, "YY-MM-DD")

        # 완료시간
        frame_time = Frame(add_tk)
        frame_time.pack(fill="x")

        lbl_time = Label(frame_time, text="완료시간", font=('맑은 고딕',12))
        lbl_time.pack(side="left")

        entry_time = Entry(frame_time, font=('맑은 고딕',12))
        entry_time.pack(side="right")
        # entry_time.insert(END, "HH:MM")

        
        # 비고
        frame_note = Frame(add_tk)
        frame_note.pack(fill="x")

        lbl_note = Label(frame_note, text="비고", font=('맑은 고딕',12))
        lbl_note.pack(side="left")

        entry_note = Entry(frame_note, font=('맑은 고딕',12))
        entry_note.pack(side="right")

        # fucntion2 frame
        frame_func2 = Frame(add_tk)
        frame_func2.pack(fill="x")

        # add db
        btn_add_db = Button(frame_func2, text="추가",command=add_db, font=('맑은 고딕',12), padx=5, pady=5)
        btn_add_db.pack(side="left")

        # cancel button
        # quit 함수를 쓰면 전체 프로그램이 종료되어서 withdraw 함수 사용함
        btn_cancel = Button(frame_func2, text="취소", command=add_tk.withdraw, font=('맑은 고딕',12), padx=5, pady=5)
        btn_cancel.pack(side="right")


        add_tk.mainloop()
    

    except Exception as err:
        msgbox.showerror("에러", err)

def get_text():
    
    selected_item = tree.item(tree.selection())
    # selected_iid = tree.focus()
    # print(selected_item)
    # print(selected_item.get("text"))
    return selected_item.get("text")

def get_iid():
    selected_iid = tree.focus()
    # print(selected_iid)
    # print(selected_item.get("text"))
    return selected_iid


# def fav_person_edit():
#     pass

def team_sort():
    df_team = pd.read_excel("C:/Python/Code/ToDoList/ToDoList_Form.xlsx", sheet_name='Team')
    # print(db_df)
    df_team = df_team.sort_values(by=['분류No'])
    
    div_team = df_team['분류'].drop_duplicates()
    div_team_cnt = len(div_team)
    team_div_list = df_team["분류"].tolist()
    print(df_team.분류No.max())

    div_team_no = 1

    if df_team.분류No.max() != div_team_cnt:
        print("다르다")

        for i in range(len(df_team['분류'])):
            if i != 0:
                if df_team['분류'][i] != df_team['분류'][i-1]:
                    div_team_no += 1    
            # else :
            df_team['분류No'][i] = div_team_no
            print(df_team['분류No'][i])


    # print(db_df)
    df_team.to_excel("C:/Python/Code/ToDoList/test.xlsx",index=False, sheet_name="Team")
    path1 = "C:/Python/Code/ToDoList/ToDoList_Form.xlsx"
    path2 = "C:/Python/Code/ToDoList/test.xlsx"
    
    wb2 = load_workbook(filename=path2)
    ws2 = wb2['Team']

    wb1 = load_workbook(filename=path1)
    wb1.remove(wb1['Team'])
    ws1 = wb1.create_sheet()
    ws1.title = 'Team'
    # ws1 = wb1['DB']
    for row in ws2:
        for cell in row:
            ws1[cell.coordinate].value = cell.value
    wb1.save(path1)


def fav_team_edit():
    fav_edit = Tk()
    fav_edit.title("팀 - 즐겨찾기 수정")


    def team_load():
        # treeview reset
        # 이게 없으면 업무 추가시 기존 데이터가 2번 돌면서 iid가 또 있다고 하면서 에러가 남
        list_team.delete(0,END)
        team_sort()
        df_team = pd.read_excel("C:/Python/Code/ToDoList/ToDoList_Form.xlsx",
                                sheet_name = 'Team')
        # print(df_team)
        # print(df_team[df_team.분류 == '전실'])
        # print(df_team['분류'][0])
        
        # print(df_team['분류'])
        div_team = df_team['분류'].drop_duplicates()
        # div_team = list(div_team)
        # div_team = df_team['분류'].drop_duplicates(ignore_index=True)
        div_team = div_team.reset_index()
        div_team = div_team.drop('index',axis=1)
        # print(div_team)
        # print(len(div_team))
        # print(div_team['분류'][1])
        # div_team = df_team[]

        for div in div_team["분류"]:
            list_team.insert(END,div)


    def get_team(self):
    # def get_team():
        # 다른 함수에서도 이 변수(team_div_value)값을 가져와야 해서 global로 형식?지정
        global team_div_value
        # 아래 문구가 있어야 list_team의 값이 다른 곳에 클릭을 해도 값이 유지가 됨
        team_div_value = list_team.get(list_team.curselection())
        # value = list_team.get(list_team.curselection())
        # print(value)

        df_team = pd.read_excel("C:/Python/Code/ToDoList/ToDoList_Form.xlsx",
                                sheet_name = 'Team')
        div_team = df_team['분류'].drop_duplicates()
        div_team = div_team.reset_index()
        div_team = div_team.drop('index',axis=1)
        # for div in div_team["분류"]:
        #     list_team.insert(END,div)

        # list_team2 reset
        list_team2.delete(0,END)
        # print(add_tk.focus())
        # print("yes")
        # df_team = pd.read_excel("C:/Python/Code/ToDoList/ToDoList_Form.xlsx",
        #                     sheet_name = 'Team')
        # print(list_team.curselection()[0])
        # print(list_team.get(list_team.curselection(),list_team.curselection()))
        # print(div_team['분류'][list_team.curselection()[0]])
        for row in range(len(df_team['분류'])):
            if df_team['분류'][row] == div_team['분류'][list_team.curselection()[0]]:
                list_team2.insert(END,df_team['이름'][row])
                # print(df_team['이름'][row])

        
        entry_team_div.delete(0,END)
        entry_team_div.insert(END, div_team['분류'][list_team.curselection()[0]])

        # print(list_team2.get(0,END)[list_team2.curselection()[0]])
        # print(list_team2.curselection()[0])
        # entry_team_name.delete(0,END)
        # entry_team_name.insert(END, df_team['이름'][list_team2.curselection()[0]])
        # print(div_team['분류'][list_team.curselection()[0]])
        # return print(div_team['분류'][list_team.curselection()[0]])
        # return team_div_value


    def get_team_name(self):
        # list_team2.get(list_team2.curselection())
        # print(get_team())
        # get_team(self)
        # print(team_div_value)
        global team_name_value
        team_name_value = list_team2.get(list_team2.curselection())

        df_team = pd.read_excel("C:/Python/Code/ToDoList/ToDoList_Form.xlsx",
                                sheet_name = 'Team')
        div_team = df_team['분류'].drop_duplicates()
        div_team = div_team.reset_index()
        div_team = div_team.drop('index',axis=1)
        
        # print(list_team2.curselection())
        # print(list_team2.curselection()[0])
        # print(list_team2.get(0,END)[list_team2.curselection()[0]])
        # print(div_team['분류'][list_team.curselection()[0]])

        # print(df_team['이름'][list_team2.curselection()[0]])
        team_name = []
        for row in range(len(df_team['분류'])):
            # print(df_team['분류'][row])
            # print()
            if df_team['분류'][row] == team_div_value:
                team_name.append(df_team['이름'][row])
                # print(df_team['이름'][row])
                # print(row)
                # print(df_team['분류'][row])

                # for row2 in range(len(df_team['이름'])):
                #     list_team2.insert(END,df_team['이름'][row])
        # print(team_name[list_team2.curselection()[0]])
        entry_team_name.delete(0,END)
        entry_team_name.insert(END,team_name[list_team2.curselection()[0]])


    def fav_add():
        
        if len(entry_team_div.get()) == 0 :
            msgbox.showerror("에러", "분류에 데이터를 입력해주세요.")
            return
        if len(entry_team_name.get()) == 0 :
            msgbox.showerror("에러", "이름에 데이터를 입력해주세요.")
            return

        df_team = pd.read_excel("C:/Python/Code/ToDoList/ToDoList_Form.xlsx",
                                sheet_name = 'Team')
        div_team = df_team['분류'].drop_duplicates()
        div_team = div_team.reset_index()
        div_team = div_team.drop('index',axis=1)


        print(len(div_team))
        div_team_cnt = len(div_team)
        wb = load_workbook("C:/Python/Code/ToDoList/ToDoList_Form.xlsx")
        team_ws = wb['Team']
            
        # Series type을 list 타입으로 변경
        # list 타입을 해야 if in (존재)를 사용할 수 있음
        team_div_list = df_team["분류"].tolist()
        # print(entry_team_div.get())
        # print(entry_team_name.get())
        # print(type(df_team['분류']))
        # print(type(team_div_list))        
        exist = 0
        for i in range(len(df_team['분류'])):
            # if df_team['분류'][row] == entry_team_div.get() :
            # print(i)
            # print(df_team['분류'][i])
            # 0부터 시작 + 1행 부터 시작 + 행 추가 : 합쳐서 3 더하기
            r_add = i + 3

            if entry_team_div.get() in team_div_list : 
                if entry_team_div.get() == df_team['분류'][i] :
                    div_team_no = df_team['분류No'][i]
                    print(entry_team_div.get())
                    print(div_team_no)
                if entry_team_name.get() == df_team['이름'][i]:
                    exist += 1
                    
                    # print(i," 분류 :", df_team['분류'][i])
                    # print(i," 이름 :  ",df_team['이름'][i])
                    # print("Yes :",df_team['이름'][i])
            else : 
                div_team_no = div_team_cnt + 1
                print(entry_team_div.get())
                print(div_team_no)   

        if exist == 0 :    
            team_ws.cell(row=r_add, column=1).value = entry_team_div.get()
            team_ws.cell(row=r_add, column=2).value = entry_team_name.get()            
            team_ws.cell(row=r_add, column=3).value = div_team_no
            msgbox.showinfo("알림", "신규 데이터가 등록되었습니다.")       
        else:
            msgbox.showerror("중복","동일한 값이 존재합니다.")

        wb.save("C:/Python/Code/ToDoList/ToDoList_Form.xlsx")

        team_load()
        

                
            
    def fav_modi():
        try:
            df_team = pd.read_excel("C:/Python/Code/ToDoList/ToDoList_Form.xlsx",
                                    sheet_name = 'Team')
            wb = load_workbook("C:/Python/Code/ToDoList/ToDoList_Form.xlsx")
            team_ws = wb['Team']
                
            # Series type을 list 타입으로 변경
            # list 타입을 해야 if in (존재)를 사용할 수 있음
            team_div_list = df_team["분류"].tolist()
            
            # print(team_div_value)
            # print(team_name_value)
            # print(team_ws.max_row)
            # if len(team_div_value) == 0 :
            #     msgbox.showerror("에러(수정)", "분류 값을 입력해주세요.")

            for r in range(2,team_ws.max_row+1) :
                if team_ws.cell(row=r,column=1).value == team_div_value:
                    if team_ws.cell(row=r,column=2).value == team_name_value:
                        team_ws.cell(row=r,column=1).value = entry_team_div.get()
                        team_ws.cell(row=r,column=2).value = entry_team_name.get()
                        # print(r, "  분류: ", entry_team_div.get())
                        # print(r, "  이름: ", entry_team_name.get())
            
            wb.save("C:/Python/Code/ToDoList/ToDoList_Form.xlsx")

            team_load()

            msgbox.showinfo("안내", "수정이 완료 되었습니다.")

        except NameError as err :
            # print("dsdfasdf")

            err_msg = str(err)
            # if "name 'team_div_value' is not defined" in err.args:
            #     print("div err")
            # logger.error(err)
            # if err_msg == "name 'team_div_value' is not defined":
            #     print("div err2")
            if 'team_name_value' in err_msg:
                for r in range(2,team_ws.max_row+1) :
                    print(team_ws.cell(row=r,column=1).value)
                    if team_ws.cell(row=r,column=1).value == team_div_value:
                        team_ws.cell(row=r,column=1).value = entry_team_div.get()

                wb.save("C:/Python/Code/ToDoList/ToDoList_Form.xlsx")
        
                team_load()

                msgbox.showinfo("안내", "수정이 완료 되었습니다.")


            else :
                # print("div err3")
                msgbox.showerror("에러", err)
            

    def fav_del():
        try:
            # print(get_iid())
            wb = load_workbook("C:/Python/Code/ToDoList/ToDoList_Form.xlsx")
            team_ws = wb['Team']

            for r in reversed(range(2,team_ws.max_row+1)):
                # print(team_ws.cell(row=r, column=1).value)
                print(team_ws.cell(row=r, column=2).value)
                if team_ws.cell(row=r,column=1).value == team_div_value:
                    if team_ws.cell(row=r,column=2).value == team_name_value:
                        team_ws.delete_rows(r)
                        print("둘다 값있고 삭제")
            
            wb.save("C:/Python/Code/ToDoList/ToDoList_Form.xlsx")
        
            team_load()

            msgbox.showinfo("안내", "삭제가 완료 되었습니다.")

                
        except NameError as err :
            err_msg = str(err)
            # if "name 'team_div_value' is not defined" in err.args:
            #     print("div err")
            # logger.error(err)
            # if err_msg == "name 'team_div_value' is not defined":
            #     print("div err2")
            if 'team_name_value' in err_msg:
                for r in reversed(range(2,team_ws.max_row+1)) :
                    print(team_ws.cell(row=r,column=1).value)
                    if team_ws.cell(row=r,column=1).value == team_div_value:
                        team_ws.delete_rows(r)
                        print("분류만 삭제")

                wb.save("C:/Python/Code/ToDoList/ToDoList_Form.xlsx")
        
                team_load()

                msgbox.showinfo("안내", "삭제가 완료 되었습니다.")


            else :
                # print("div err3")
                msgbox.showerror("에러", err)



    # 분류 프레임
    frame_team_div = Frame(fav_edit)
    frame_team_div.pack(fill="x")

    lbl_team_div = Label(frame_team_div, text="분류")
    lbl_team_div.pack(side="left", pady=3)

    entry_team_div = Entry(frame_team_div)
    entry_team_div.pack(side="right")

    # 팀 이름 프레임
    frame_team_name = Frame(fav_edit)
    frame_team_name.pack(fill="x")

    lbl_team_name = Label(frame_team_name, text="이름")
    lbl_team_name.pack(side="left", pady=3)

    entry_team_name = Entry(frame_team_name)
    entry_team_name.pack(side="right")



    # 담당팀(복수)
    frame_team_name2 = Frame(fav_edit)
    # frame_team_name2.pack()
    frame_team_name2.pack(fill="x", pady=3)

    lbl_team_name2 = Label(frame_team_name2, text="2) 담당팀(다수)")
    lbl_team_name2.pack(side="left")


    # 담당팀(복수) - 리스트
    frame_team_name3 = Frame(fav_edit)
    # frame_team_name3.pack()
    # frame_team_name3.pack(side="left")
    frame_team_name3.pack(fill="both")
    # frame_team_name3.pack(fill="x")

    yscrollbar = Scrollbar(frame_team_name3)
    yscrollbar.pack(side="left", fill='y')

    # xscrollbar x축의 절반만 나오게 하는 것 실패
    xscrollbar = Scrollbar(frame_team_name3, orient=HORIZONTAL)
    # xscrollbar.pack(side="bottom")
    # xscrollbar.pack(side="bottom",expand=True)
    # xscrollbar.pack(side="bottom", fill='x',expand=True)
    xscrollbar.pack(side="bottom", fill='x')
    
    # xscrollbar.place(relwidth=0.5)

    list_team = Listbox(frame_team_name3, selectmode="browse", height = 6,
    yscrollcommand=yscrollbar.set, xscrollcommand=xscrollbar.set)

    list_team.bind('<<ListboxSelect>>',get_team)        

    # list_team.pack(side="left", expand=1)
    # list_team.place(x=0, y=200)
    # list_team.pack(side="left", fill="y")
    list_team.pack(side="left", fill="both", expand=True)


    yscrollbar.config(command=list_team.yview)
    xscrollbar.config(command=list_team.xview)

    # get_team()

    # 담당팀(복수) - 리스트
    # frame_team_name5 = Frame(add_tk)
    # frame_team_name5.pack()
    # frame_team_name5.pack(side="right")

    yscrollbar2 = Scrollbar(frame_team_name3)
    yscrollbar2.pack(side="right", fill='y')

    xscrollbar2 = Scrollbar(frame_team_name3, orient=HORIZONTAL)
    # xscrollbar2.pack(side="bottom")
    xscrollbar2.pack(side="bottom", fill='x',expand=True)

    list_team2 = Listbox(frame_team_name3, selectmode="extended", height = 6,
    yscrollcommand=yscrollbar2.set, xscrollcommand=xscrollbar2.set)

    # list_file = Listbox(list_frame, selectmode="extended", height = 10,
    # yscrollcommand=yscrollbar.set, xscrollcommand=xscrollbar.set, wrap=NONE)
    # list_team2.pack(side="left", expand=1)
    list_team2.pack(side="right", fill="both", expand=True)


    yscrollbar2.config(command=list_team2.yview)
    xscrollbar2.config(command=list_team2.xview)

    list_team2.bind('<<ListboxSelect>>',get_team_name)        

    team_load()

    # fucntion frame
    frame_func = Frame(fav_edit)
    frame_func.pack(fill="x", pady=3)

    btn_fav_add = Button(frame_func, text="추가", command=fav_add)
    btn_fav_add.pack(side="left", pady=3, padx=3)
    
    btn_fav_modi = Button(frame_func, text="수정", command=fav_modi)
    btn_fav_modi.pack(side="left", pady=3, padx=3)
    
    btn_fav_del = Button(frame_func, text="삭제", command=fav_del)
    btn_fav_del.pack(side="left", pady=3, padx=3)


    fav_edit.mainloop()

# 즐겨찾기 담당자 정렬

def person_sort():
    df_person = pd.read_excel("C:/Python/Code/ToDoList/ToDoList_Form.xlsx", sheet_name='Person')
    # print(db_df)
    df_person = df_person.sort_values(by=['분류No'])
    
    div_team = df_person['분류'].drop_duplicates()
    div_team_cnt = len(div_team)
    person_div_list = df_person["분류"].tolist()
    print(df_person.분류No.max())

    div_team_no = 1

    if df_person.분류No.max() != div_team_cnt:
        print("다르다")

        for i in range(len(df_person['분류'])):
            if i != 0:
                if df_person['분류'][i] != df_person['분류'][i-1]:
                    div_team_no += 1    
            # else :
            df_person['분류No'][i] = div_team_no
            print(df_person['분류No'][i])


    # print(db_df)
    df_person.to_excel("C:/Python/Code/ToDoList/test.xlsx",index=False, sheet_name="Person")
    path1 = "C:/Python/Code/ToDoList/ToDoList_Form.xlsx"
    path2 = "C:/Python/Code/ToDoList/test.xlsx"
    
    wb2 = load_workbook(filename=path2)
    ws2 = wb2['Person']

    wb1 = load_workbook(filename=path1)
    wb1.remove(wb1['Person'])
    ws1 = wb1.create_sheet()
    ws1.title = 'Person'
    # ws1 = wb1['DB']
    for row in ws2:
        for cell in row:
            ws1[cell.coordinate].value = cell.value
    wb1.save(path1)




# 즐겨찾기 담당자 편집

def fav_person_edit():
    fav_person_edit = Tk()
    fav_person_edit.title("담당자 - 즐겨찾기 수정")


    def person_load():
        # treeview reset
        # 이게 없으면 업무 추가시 기존 데이터가 2번 돌면서 iid가 또 있다고 하면서 에러가 남
        list_person.delete(0,END)
        person_sort()
        df_person = pd.read_excel("C:/Python/Code/ToDoList/ToDoList_Form.xlsx",
                                sheet_name = 'Person')
        # print(df_person)
        # print(df_person[df_person.분류 == '전실'])
        # print(df_person['분류'][0])
        
        # print(df_person['분류'])
        div_person = df_person['분류'].drop_duplicates()
        # div_person = list(div_person)
        # div_person = df_person['분류'].drop_duplicates(ignore_index=True)
        div_person = div_person.reset_index()
        div_person = div_person.drop('index',axis=1)
        # print(div_person)
        # print(len(div_person))
        # print(div_person['분류'][1])
        # div_person = df_person[]

        for div in div_person["분류"]:
            list_person.insert(END,div)


    def get_person(self):
    # def get_person():
        # 다른 함수에서도 이 변수(person_div_value)값을 가져와야 해서 global로 형식?지정
        global person_div_value
        # 아래 문구가 있어야 list_person의 값이 다른 곳에 클릭을 해도 값이 유지가 됨
        person_div_value = list_person.get(list_person.curselection())
        # value = list_person.get(list_person.curselection())
        # print(value)

        df_person = pd.read_excel("C:/Python/Code/ToDoList/ToDoList_Form.xlsx",
                                sheet_name = 'Person')
        div_person = df_person['분류'].drop_duplicates()
        div_person = div_person.reset_index()
        div_person = div_person.drop('index',axis=1)
        # for div in div_person["분류"]:
        #     list_person.insert(END,div)

        # list_person2 reset
        list_person2.delete(0,END)
        # print(add_tk.focus())
        # print("yes")
        # df_person = pd.read_excel("C:/Python/Code/ToDoList/ToDoList_Form.xlsx",
        #                     sheet_name = 'person')
        # print(list_person.curselection()[0])
        # print(list_person.get(list_person.curselection(),list_person.curselection()))
        # print(div_person['분류'][list_person.curselection()[0]])
        for row in range(len(df_person['분류'])):
            if df_person['분류'][row] == div_person['분류'][list_person.curselection()[0]]:
                list_person2.insert(END,df_person['이름'][row])
                # print(df_person['이름'][row])

        
        entry_person_div.delete(0,END)
        entry_person_div.insert(END, div_person['분류'][list_person.curselection()[0]])

        # print(list_person2.get(0,END)[list_person2.curselection()[0]])
        # print(list_person2.curselection()[0])
        # entry_person_name.delete(0,END)
        # entry_person_name.insert(END, df_person['이름'][list_person2.curselection()[0]])
        # print(div_person['분류'][list_person.curselection()[0]])
        # return print(div_person['분류'][list_person.curselection()[0]])
        # return person_div_value


    def get_person_name(self):
        # list_person2.get(list_person2.curselection())
        # print(get_person())
        # get_person(self)
        # print(person_div_value)
        global person_name_value
        person_name_value = list_person2.get(list_person2.curselection())

        df_person = pd.read_excel("C:/Python/Code/ToDoList/ToDoList_Form.xlsx",
                                sheet_name = 'Person')
        div_person = df_person['분류'].drop_duplicates()
        div_person = div_person.reset_index()
        div_person = div_person.drop('index',axis=1)
        
        # print(list_person2.curselection())
        # print(list_person2.curselection()[0])
        # print(list_person2.get(0,END)[list_person2.curselection()[0]])
        # print(div_person['분류'][list_person.curselection()[0]])

        # print(df_person['이름'][list_person2.curselection()[0]])
        person_name = []
        for row in range(len(df_person['분류'])):
            # print(df_person['분류'][row])
            # print()
            if df_person['분류'][row] == person_div_value:
                person_name.append(df_person['이름'][row])
                # print(df_person['이름'][row])
                # print(row)
                # print(df_person['분류'][row])

                # for row2 in range(len(df_person['이름'])):
                #     list_person2.insert(END,df_person['이름'][row])
        # print(person_name[list_person2.curselection()[0]])
        entry_person_name.delete(0,END)
        entry_person_name.insert(END,person_name[list_person2.curselection()[0]])


    def fav_add():
        
        if len(entry_person_div.get()) == 0 :
            msgbox.showerror("에러", "분류에 데이터를 입력해주세요.")
            return
        if len(entry_person_name.get()) == 0 :
            msgbox.showerror("에러", "이름에 데이터를 입력해주세요.")
            return

        df_person = pd.read_excel("C:/Python/Code/ToDoList/ToDoList_Form.xlsx",
                                sheet_name = 'Person')
        div_person = df_person['분류'].drop_duplicates()
        div_person = div_person.reset_index()
        div_person = div_person.drop('index',axis=1)


        print(len(div_person))
        div_person_cnt = len(div_person)
        wb = load_workbook("C:/Python/Code/ToDoList/ToDoList_Form.xlsx")
        person_ws = wb['Person']
            
        # Series type을 list 타입으로 변경
        # list 타입을 해야 if in (존재)를 사용할 수 있음
        person_div_list = df_person["분류"].tolist()
        # print(entry_person_div.get())
        # print(entry_person_name.get())
        # print(type(df_person['분류']))
        # print(type(person_div_list))        
        exist = 0
        for i in range(len(df_person['분류'])):
            # if df_person['분류'][row] == entry_person_div.get() :
            # print(i)
            # print(df_person['분류'][i])
            # 0부터 시작 + 1행 부터 시작 + 행 추가 : 합쳐서 3 더하기
            r_add = i + 3

            if entry_person_div.get() in person_div_list : 
                if entry_person_div.get() == df_person['분류'][i] :
                    div_person_no = df_person['분류No'][i]
                    print(entry_person_div.get())
                    print(div_person_no)
                if entry_person_name.get() == df_person['이름'][i]:
                    exist += 1
                    
                    # print(i," 분류 :", df_person['분류'][i])
                    # print(i," 이름 :  ",df_person['이름'][i])
                    # print("Yes :",df_person['이름'][i])
            else : 
                div_person_no = div_person_cnt + 1
                print(entry_person_div.get())
                print(div_person_no)   

        if exist == 0 :    
            person_ws.cell(row=r_add, column=1).value = entry_person_div.get()
            person_ws.cell(row=r_add, column=2).value = entry_person_name.get()            
            person_ws.cell(row=r_add, column=3).value = div_person_no
            msgbox.showinfo("알림", "신규 데이터가 등록되었습니다.")       
        else:
            msgbox.showerror("중복","동일한 값이 존재합니다.")

        wb.save("C:/Python/Code/ToDoList/ToDoList_Form.xlsx")

        person_load()
        

                
            
    def fav_modi():
        try:
            df_person = pd.read_excel("C:/Python/Code/ToDoList/ToDoList_Form.xlsx",
                                    sheet_name = 'Person')
            wb = load_workbook("C:/Python/Code/ToDoList/ToDoList_Form.xlsx")
            person_ws = wb['Person']
                
            # Series type을 list 타입으로 변경
            # list 타입을 해야 if in (존재)를 사용할 수 있음
            person_div_list = df_person["분류"].tolist()
            
            # print(person_div_value)
            # print(person_name_value)
            # print(person_ws.max_row)
            # if len(person_div_value) == 0 :
            #     msgbox.showerror("에러(수정)", "분류 값을 입력해주세요.")

            for r in range(2,person_ws.max_row+1) :
                if person_ws.cell(row=r,column=1).value == person_div_value:
                    if person_ws.cell(row=r,column=2).value == person_name_value:
                        person_ws.cell(row=r,column=1).value = entry_person_div.get()
                        person_ws.cell(row=r,column=2).value = entry_person_name.get()
                        # print(r, "  분류: ", entry_person_div.get())
                        # print(r, "  이름: ", entry_person_name.get())
            
            wb.save("C:/Python/Code/ToDoList/ToDoList_Form.xlsx")

            person_load()

            msgbox.showinfo("안내", "수정이 완료 되었습니다.")

        except NameError as err :
            # print("dsdfasdf")

            err_msg = str(err)
            # if "name 'person_div_value' is not defined" in err.args:
            #     print("div err")
            # logger.error(err)
            # if err_msg == "name 'person_div_value' is not defined":
            #     print("div err2")
            if 'person_name_value' in err_msg:
                for r in range(2,person_ws.max_row+1) :
                    print(person_ws.cell(row=r,column=1).value)
                    if person_ws.cell(row=r,column=1).value == person_div_value:
                        person_ws.cell(row=r,column=1).value = entry_person_div.get()

                wb.save("C:/Python/Code/ToDoList/ToDoList_Form.xlsx")
        
                person_load()

                msgbox.showinfo("안내", "수정이 완료 되었습니다.")


            else :
                # print("div err3")
                msgbox.showerror("에러", err)
            

    def fav_del():
        try:
            # print(get_iid())
            wb = load_workbook("C:/Python/Code/ToDoList/ToDoList_Form.xlsx")
            person_ws = wb['Person']

            for r in reversed(range(2,person_ws.max_row+1)):
                # print(person_ws.cell(row=r, column=1).value)
                print(person_ws.cell(row=r, column=2).value)
                if person_ws.cell(row=r,column=1).value == person_div_value:
                    if person_ws.cell(row=r,column=2).value == person_name_value:
                        person_ws.delete_rows(r)
                        print("둘다 값있고 삭제")
            
            wb.save("C:/Python/Code/ToDoList/ToDoList_Form.xlsx")
        
            person_load()

            msgbox.showinfo("안내", "삭제가 완료 되었습니다.")

                
        except NameError as err :
            err_msg = str(err)
            # if "name 'person_div_value' is not defined" in err.args:
            #     print("div err")
            # logger.error(err)
            # if err_msg == "name 'person_div_value' is not defined":
            #     print("div err2")
            if 'person_name_value' in err_msg:
                for r in reversed(range(2,person_ws.max_row+1)) :
                    print(person_ws.cell(row=r,column=1).value)
                    if person_ws.cell(row=r,column=1).value == person_div_value:
                        person_ws.delete_rows(r)
                        print("분류만 삭제")

                wb.save("C:/Python/Code/ToDoList/ToDoList_Form.xlsx")
        
                person_load()

                msgbox.showinfo("안내", "삭제가 완료 되었습니다.")


            else :
                # print("div err3")
                msgbox.showerror("에러", err)



    # 분류 프레임
    frame_person_div = Frame(fav_person_edit)
    frame_person_div.pack(fill="x")

    lbl_person_div = Label(frame_person_div, text="분류")
    lbl_person_div.pack(side="left", pady=3)

    entry_person_div = Entry(frame_person_div)
    entry_person_div.pack(side="right")

    # 팀 이름 프레임
    frame_person_name = Frame(fav_person_edit)
    frame_person_name.pack(fill="x")

    lbl_person_name = Label(frame_person_name, text="이름")
    lbl_person_name.pack(side="left", pady=3)

    entry_person_name = Entry(frame_person_name)
    entry_person_name.pack(side="right")



    # 담당팀(복수)
    frame_person_name2 = Frame(fav_person_edit)
    # frame_person_name2.pack()
    frame_person_name2.pack(fill="x", pady=3)

    lbl_person_name2 = Label(frame_person_name2, text="2) 담당팀(다수)")
    lbl_person_name2.pack(side="left")


    # 담당팀(복수) - 리스트
    frame_person_name3 = Frame(fav_person_edit)
    # frame_person_name3.pack()
    # frame_person_name3.pack(side="left")
    frame_person_name3.pack(fill="both")
    # frame_person_name3.pack(fill="x")

    yscrollbar = Scrollbar(frame_person_name3)
    yscrollbar.pack(side="left", fill='y')

    # xscrollbar x축의 절반만 나오게 하는 것 실패
    xscrollbar = Scrollbar(frame_person_name3, orient=HORIZONTAL)
    # xscrollbar.pack(side="bottom")
    # xscrollbar.pack(side="bottom",expand=True)
    # xscrollbar.pack(side="bottom", fill='x',expand=True)
    xscrollbar.pack(side="bottom", fill='x')
    
    # xscrollbar.place(relwidth=0.5)

    list_person = Listbox(frame_person_name3, selectmode="browse", height = 6,
    yscrollcommand=yscrollbar.set, xscrollcommand=xscrollbar.set)

    list_person.bind('<<ListboxSelect>>',get_person)        

    # list_person.pack(side="left", expand=1)
    # list_person.place(x=0, y=200)
    # list_person.pack(side="left", fill="y")
    list_person.pack(side="left", fill="both", expand=True)


    yscrollbar.config(command=list_person.yview)
    xscrollbar.config(command=list_person.xview)

    # get_person()

    # 담당팀(복수) - 리스트
    # frame_person_name5 = Frame(add_tk)
    # frame_person_name5.pack()
    # frame_person_name5.pack(side="right")

    yscrollbar2 = Scrollbar(frame_person_name3)
    yscrollbar2.pack(side="right", fill='y')

    xscrollbar2 = Scrollbar(frame_person_name3, orient=HORIZONTAL)
    # xscrollbar2.pack(side="bottom")
    xscrollbar2.pack(side="bottom", fill='x',expand=True)

    list_person2 = Listbox(frame_person_name3, selectmode="extended", height = 6,
    yscrollcommand=yscrollbar2.set, xscrollcommand=xscrollbar2.set)

    # list_file = Listbox(list_frame, selectmode="extended", height = 10,
    # yscrollcommand=yscrollbar.set, xscrollcommand=xscrollbar.set, wrap=NONE)
    # list_person2.pack(side="left", expand=1)
    list_person2.pack(side="right", fill="both", expand=True)


    yscrollbar2.config(command=list_person2.yview)
    xscrollbar2.config(command=list_person2.xview)

    list_person2.bind('<<ListboxSelect>>',get_person_name)        

    person_load()

    # fucntion frame
    frame_func = Frame(fav_person_edit)
    frame_func.pack(fill="x", pady=3)

    btn_fav_add = Button(frame_func, text="추가", command=fav_add)
    btn_fav_add.pack(side="left", pady=3, padx=3)
    
    btn_fav_modi = Button(frame_func, text="수정", command=fav_modi)
    btn_fav_modi.pack(side="left", pady=3, padx=3)
    
    btn_fav_del = Button(frame_func, text="삭제", command=fav_del)
    btn_fav_del.pack(side="left", pady=3, padx=3)


    fav_person_edit.mainloop()



def add_task2():

    # print(get_iid()[6:8])
    # Lv3 업무 선택 후 하위 업무를 선택하면 에러 메시지 구현
    if get_iid()[6:8] != '00':
        msgbox.showerror("에러-하위 업무 추가", "최하위단 업무를 선택하셨습니다.\n\n 상위 업무를 선택해주세요.")
        return

    try:
        add_tk = Tk()
        add_tk.title("하위 업무 추가")


        def add_db():
            wb = load_workbook("C:/Python/Code/ToDoList/ToDoList_Form.xlsx")
            db_ws = wb['DB']
            # values_only=True를 해야 위치값이 아닌 실제 데이터 값이 도출됨
            db_header = db_ws.iter_rows(min_row=1, max_row=1, max_col=9, values_only=True)
            db_datas = db_ws.iter_rows(min_row=2, max_col=9, values_only=True)
            db_header = [r for r in db_header]
            db_datas = [r for r in db_datas]
            
            # 선택 항목의 iid값 가져오기
            for i in range(1, db_ws.max_row+1):
                if db_ws.cell(i,1).value == get_iid():
                    # print(i)
                    r_selected = i
            # print(db_datas[r_selected-2][0][3:8])
            # print(db_datas[r_selected-2][0])        
            
            lv1_cnt = 0
            lv2_cnt = 0
            lv3_cnt = 0
            print(db_ws.max_row)
            for x in range(db_ws.max_row-1):
                # print(db_datas[x][0][0:2])
                # 선택 값이 1레벨 -> 하위 업무는 2레벨에 업무 추가
                if db_datas[r_selected-2][0][3:8] == "00-00":
                    # print(db_datas[r_selected-2][0])
                    print(x)        
                    print(db_datas[x][0][0:2])
                    print(db_datas[r_selected-2][0][0:2])
                    if db_datas[x][0][0:2] == db_datas[r_selected-2][0][0:2]:
                        lv1_cnt = int(db_datas[x][0][0:2])
                        lv2_cnt = int(db_datas[x][0][3:5])
                        print(lv1_cnt)
                # 선택 값이 2레벨 -> 하위 업무는 3레벨에 업무 추가
                elif db_datas[r_selected-2][0][6:8] == "00":
                    if db_datas[x][0][0:5] == db_datas[r_selected-2][0][0:5]:
                        lv1_cnt = int(db_datas[x][0][0:2])
                        lv2_cnt = int(db_datas[x][0][3:5])
                        lv3_cnt = int(db_datas[x][0][6:8])
            # print(lv1_cnt)
            # 다음 iid 값 넣기        
            lv1_cnt = str(format(lv1_cnt,'02'))
            print(lv1_cnt)
            
            lv2_next = lv2_cnt+1
            lv2_next = format(lv2_cnt+1,'02')
            lv2_cnt = str(format(lv2_cnt,'02'))
            
            lv3_next = lv3_cnt+1
            lv3_next = format(lv3_cnt+1,'02')
            lv3_cnt = str(format(lv3_cnt,'02'))
            
            # print(rad_var.get())
            
            # print(lv1_cnt + "-" + lv2_next + "-00")
            # db_datas는 0부터 시작하고, 제목행이 빠져서 -2를 해야함
            # 선택 값이 1레벨 -> 하위 업무는 2레벨에 업무 추가
            if db_datas[r_selected-2][0][3:8] == "00-00":
                # 단일팀/사람
                if rad_var.get() == 1 :
                    # print("rad_var == 1")
                    r_add = db_ws.max_row + 1
                    db_ws.cell(row=r_add, column=1).value = lv1_cnt + "-" + lv2_next + "-00"
                    if len(entry_task_name.get()) == 0:
                        db_ws.cell(row=r_add, column=2).value = " "
                    else :    
                        db_ws.cell(row=r_add, column=2).value = entry_task_name.get()
                    if len(entry_team_name.get()) == 0 :
                        db_ws.cell(row=r_add, column=3).value = " "
                    else :    
                        db_ws.cell(row=r_add, column=3).value = entry_team_name.get()
                    if len(entry_person_name.get()) == 0:
                        db_ws.cell(row=r_add, column=4).value = " "
                    else :    
                        db_ws.cell(row=r_add, column=4).value = entry_person_name.get()
                    db_ws.cell(row=r_add, column=5).value = cmb_situation.get()
                    if len(entry_date.get()) == 0:
                        db_ws.cell(row=r_add, column=6).value = " "    
                    else :
                        db_ws.cell(row=r_add, column=6).value = entry_date.get()
                    if len(entry_time.get()) == 0:
                        db_ws.cell(row=r_add, column=7).value = " "
                    else :    
                        db_ws.cell(row=r_add, column=7).value = entry_time.get()
                    if len(entry_note.get()) == 0:
                        db_ws.cell(row=r_add, column=8).value = " "
                    else :     
                        db_ws.cell(row=r_add, column=8).value = entry_note.get()
                    db_ws.cell(row=r_add, column=9).value = 2

                    wb.save("C:/Python/Code/ToDoList/ToDoList_Form.xlsx")
                # 다수 팀
                elif rad_var.get() == 2 :
                    # print("rad_var==2")
                    print(lv1_cnt)
                    file_cnt = list_team2.size()
                    if file_cnt > 0 :
                        for j in range(file_cnt):
                            # print(list_team2.get(0,END)[j])
                            # print(lv2_next)
                            r_add = db_ws.max_row + 1
                            db_ws.cell(row=r_add, column=1).value = lv1_cnt + "-" + lv2_next + "-00"
                            print(db_ws.cell(row=r_add, column=1).value)
                            if len(entry_task_name.get()) == 0:
                                db_ws.cell(row=r_add, column=2).value = " "
                            else :    
                                db_ws.cell(row=r_add, column=2).value = entry_task_name.get()
                            # if len(entry_team_name.get()) == 0 :
                                # return
                                # db_ws.cell(row=r_add, column=3).value = " "
                            # else :    
                            db_ws.cell(row=r_add, column=3).value = list_team2.get(0,END)[j]
                            
                            if len(entry_person_name.get()) == 0:
                                db_ws.cell(row=r_add, column=4).value = " "
                            else :    
                                db_ws.cell(row=r_add, column=4).value = entry_person_name.get()
                            db_ws.cell(row=r_add, column=5).value = cmb_situation.get()
                            if len(entry_date.get()) == 0:
                                db_ws.cell(row=r_add, column=6).value = " "    
                            else :
                                db_ws.cell(row=r_add, column=6).value = entry_date.get()
                            if len(entry_time.get()) == 0:
                                db_ws.cell(row=r_add, column=7).value = " "
                            else :    
                                db_ws.cell(row=r_add, column=7).value = entry_time.get()
                            if len(entry_note.get()) == 0:
                                db_ws.cell(row=r_add, column=8).value = " "
                            else :     
                                db_ws.cell(row=r_add, column=8).value = entry_note.get()
                            db_ws.cell(row=r_add, column=9).value = 2

                            wb.save("C:/Python/Code/ToDoList/ToDoList_Form.xlsx")

                            lv2_next = int(lv2_next)
                            lv2_next = format(lv2_next+1,'02')

                # 다수 담당자
                elif rad_var.get() == 3 :
                    # print("rad_var==2")
                    # print(lv1_cnt)
                    file_cnt = list_person2.size()
                    if file_cnt > 0 :
                        for j in range(file_cnt):
                            # print(list_team2.get(0,END)[j])
                            # print(lv2_next)
                            r_add = db_ws.max_row + 1
                            db_ws.cell(row=r_add, column=1).value = lv1_cnt + "-" + lv2_next + "-00"
                            print(db_ws.cell(row=r_add, column=1).value)
                            if len(entry_task_name.get()) == 0:
                                db_ws.cell(row=r_add, column=2).value = " "
                            else :    
                                db_ws.cell(row=r_add, column=2).value = entry_task_name.get()
                            if len(entry_team_name.get()) == 0 :
                                db_ws.cell(row=r_add, column=3).value = " "
                            else :    
                                db_ws.cell(row=r_add, column=3).value = entry_team_name.get()
                            
                            # if len(entry_person_name.get()) == 0:
                            #     db_ws.cell(row=r_add, column=4).value = " "
                            # else :    
                            #     db_ws.cell(row=r_add, column=4).value = entry_person_name.get()
                            
                            db_ws.cell(row=r_add, column=4).value = list_person2.get(0,END)[j]                            

                            db_ws.cell(row=r_add, column=5).value = cmb_situation.get()
                            if len(entry_date.get()) == 0:
                                db_ws.cell(row=r_add, column=6).value = " "    
                            else :
                                db_ws.cell(row=r_add, column=6).value = entry_date.get()
                            if len(entry_time.get()) == 0:
                                db_ws.cell(row=r_add, column=7).value = " "
                            else :    
                                db_ws.cell(row=r_add, column=7).value = entry_time.get()
                            if len(entry_note.get()) == 0:
                                db_ws.cell(row=r_add, column=8).value = " "
                            else :     
                                db_ws.cell(row=r_add, column=8).value = entry_note.get()
                            db_ws.cell(row=r_add, column=9).value = 2

                            wb.save("C:/Python/Code/ToDoList/ToDoList_Form.xlsx")

                            lv2_next = int(lv2_next)
                            lv2_next = format(lv2_next+1,'02')

                        
            # 선택 값이 2레벨 -> 하위 업무는 3레벨에 업무 추가
            elif db_datas[r_selected-2][0][6:8] == "00":
                # 단일팀/사람
                if rad_var.get() == 1 :
                    r_add = db_ws.max_row + 1
                    db_ws.cell(row=r_add, column=1).value = lv1_cnt + "-" + lv2_cnt + "-" + lv3_next

                    if len(entry_task_name.get()) == 0:
                        db_ws.cell(row=r_add, column=2).value = " "
                    else :    
                        db_ws.cell(row=r_add, column=2).value = entry_task_name.get()
                    if len(entry_team_name.get()) == 0 :
                        db_ws.cell(row=r_add, column=3).value = " "
                    else :    
                        db_ws.cell(row=r_add, column=3).value = entry_team_name.get()
                    if len(entry_person_name.get()) == 0:
                        db_ws.cell(row=r_add, column=4).value = " "
                    else :    
                        db_ws.cell(row=r_add, column=4).value = entry_person_name.get()
                    db_ws.cell(row=r_add, column=5).value = cmb_situation.get()
                    if len(entry_date.get()) == 0:
                        db_ws.cell(row=r_add, column=6).value = " "    
                    else :
                        db_ws.cell(row=r_add, column=6).value = entry_date.get()
                    if len(entry_time.get()) == 0:
                        db_ws.cell(row=r_add, column=7).value = " "
                    else :    
                        db_ws.cell(row=r_add, column=7).value = entry_time.get()
                    if len(entry_note.get()) == 0:
                        db_ws.cell(row=r_add, column=8).value = " "
                    else :     
                        db_ws.cell(row=r_add, column=8).value = entry_note.get()
                    db_ws.cell(row=r_add, column=9).value = 3

                    wb.save("C:/Python/Code/ToDoList/ToDoList_Form.xlsx")

                # 다수 팀
                elif rad_var.get() == 2 :
                    # print("rad_var==2")
                    # print(lv1_cnt)
                    file_cnt = list_team2.size()
                    if file_cnt > 0 :
                        for j in range(file_cnt):
                            # print(list_team2.get(0,END)[j])
                            # print(lv2_next)
                            r_add = db_ws.max_row + 1
                            db_ws.cell(row=r_add, column=1).value = lv1_cnt + "-" + lv2_cnt + "-" + lv3_next
                            if len(entry_task_name.get()) == 0:
                                db_ws.cell(row=r_add, column=2).value = " "
                            else :    
                                db_ws.cell(row=r_add, column=2).value = entry_task_name.get()
                            # if len(entry_team_name.get()) == 0 :
                                # return
                                # db_ws.cell(row=r_add, column=3).value = " "
                            # else :    
                            db_ws.cell(row=r_add, column=3).value = list_team2.get(0,END)[j]
                            
                            if len(entry_person_name.get()) == 0:
                                db_ws.cell(row=r_add, column=4).value = " "
                            else :    
                                db_ws.cell(row=r_add, column=4).value = entry_person_name.get()
                            db_ws.cell(row=r_add, column=5).value = cmb_situation.get()
                            if len(entry_date.get()) == 0:
                                db_ws.cell(row=r_add, column=6).value = " "    
                            else :
                                db_ws.cell(row=r_add, column=6).value = entry_date.get()
                            if len(entry_time.get()) == 0:
                                db_ws.cell(row=r_add, column=7).value = " "
                            else :    
                                db_ws.cell(row=r_add, column=7).value = entry_time.get()
                            if len(entry_note.get()) == 0:
                                db_ws.cell(row=r_add, column=8).value = " "
                            else :     
                                db_ws.cell(row=r_add, column=8).value = entry_note.get()
                            db_ws.cell(row=r_add, column=9).value = 2

                            wb.save("C:/Python/Code/ToDoList/ToDoList_Form.xlsx")

                            lv3_next = int(lv3_next)
                            lv3_next = format(lv3_next+1,'02')

                # 다수 담당자
                elif rad_var.get() == 3 :
                    # print("rad_var==2")
                    # print(lv1_cnt)
                    file_cnt = list_person2.size()
                    if file_cnt > 0 :
                        for j in range(file_cnt):
                            # print(list_team2.get(0,END)[j])
                            # print(lv2_next)
                            r_add = db_ws.max_row + 1
                            db_ws.cell(row=r_add, column=1).value = lv1_cnt + "-" + lv2_cnt + "-" + lv3_next
                            print(db_ws.cell(row=r_add, column=1).value)
                            if len(entry_task_name.get()) == 0:
                                db_ws.cell(row=r_add, column=2).value = " "
                            else :    
                                db_ws.cell(row=r_add, column=2).value = entry_task_name.get()
                            if len(entry_team_name.get()) == 0 :
                                db_ws.cell(row=r_add, column=3).value = " "
                            else :    
                                db_ws.cell(row=r_add, column=3).value = entry_team_name.get()
                            
                            # if len(entry_person_name.get()) == 0:
                            #     db_ws.cell(row=r_add, column=4).value = " "
                            # else :    
                            #     db_ws.cell(row=r_add, column=4).value = entry_person_name.get()
                            
                            db_ws.cell(row=r_add, column=4).value = list_person2.get(0,END)[j]                            

                            db_ws.cell(row=r_add, column=5).value = cmb_situation.get()
                            if len(entry_date.get()) == 0:
                                db_ws.cell(row=r_add, column=6).value = " "    
                            else :
                                db_ws.cell(row=r_add, column=6).value = entry_date.get()
                            if len(entry_time.get()) == 0:
                                db_ws.cell(row=r_add, column=7).value = " "
                            else :    
                                db_ws.cell(row=r_add, column=7).value = entry_time.get()
                            if len(entry_note.get()) == 0:
                                db_ws.cell(row=r_add, column=8).value = " "
                            else :     
                                db_ws.cell(row=r_add, column=8).value = entry_note.get()
                            db_ws.cell(row=r_add, column=9).value = 2

                            wb.save("C:/Python/Code/ToDoList/ToDoList_Form.xlsx")

                            lv3_next = int(lv3_next)
                            lv3_next = format(lv3_next+1,'02')


            msgbox.showinfo("알림", "하위 업무 추가가 완료되었습니다.")

            add_tk.withdraw()
            sort()
            load_task()


        def team_load():
            df_team = pd.read_excel("C:/Python/Code/ToDoList/ToDoList_Form.xlsx",
                                    sheet_name = 'Team')
            # print(df_team)
            # print(df_team[df_team.분류 == '전실'])
            # print(df_team['분류'][0])
            
            # print(df_team['분류'])
            div_team = df_team['분류'].drop_duplicates()
            # div_team = list(div_team)
            # div_team = df_team['분류'].drop_duplicates(ignore_index=True)
            div_team = div_team.reset_index()
            div_team = div_team.drop('index',axis=1)
            # print(div_team)
            # print(len(div_team))
            # print(div_team['분류'][1])
            # div_team = df_team[]

            for div in div_team["분류"]:
                list_team.insert(END,div)

        def person_load():
            df_person = pd.read_excel("C:/Python/Code/ToDoList/ToDoList_Form.xlsx",
                                    sheet_name = 'Person')

            div_person = df_person['분류'].drop_duplicates()
            div_person = div_person.reset_index()
            div_person = div_person.drop('index',axis=1)

            for div in div_person["분류"]:
                list_person.insert(END,div)

        def get_team(self):
            # 아래 문구가 있어야 list_team의 값이 다른 곳에 클릭을 해도 값이 유지가 됨
            list_team.get(list_team.curselection())
            # value = list_team.get(list_team.curselection())
            # print(value)

            df_team = pd.read_excel("C:/Python/Code/ToDoList/ToDoList_Form.xlsx",
                                    sheet_name = 'Team')
            div_team = df_team['분류'].drop_duplicates()
            div_team = div_team.reset_index()
            div_team = div_team.drop('index',axis=1)
            # for div in div_team["분류"]:
            #     list_team.insert(END,div)

            # list_team2 reset
            list_team2.delete(0,END)
            # print(add_tk.focus())
            # print("yes")
            # df_team = pd.read_excel("C:/Python/Code/ToDoList/ToDoList_Form.xlsx",
            #                     sheet_name = 'Team')
            # print(list_team.curselection()[0])
            # print(list_team.get(list_team.curselection(),list_team.curselection()))
            # print(div_team['분류'][list_team.curselection()[0]])
            for row in range(len(df_team['분류'])):
                if df_team['분류'][row] == div_team['분류'][list_team.curselection()[0]]:
                    list_team2.insert(END,df_team['이름'][row])
            
            # return list_team2.get(0,END)

            ######bbox 사용해서 listbox 일때만 클릭되게 해보기?

            # add_tk.bind('<Button-1>',get_team2)
            # add_tk.bind('<Button-1>',get_team)
            # add_tk.bind('<ButtonRelease-1>',get_team)

            # for row in df_team:
            #     if df_team['분류'][row] == list_team.get():
            #         list_team2.insert(END,df_team['이름'][row])

        def get_person(self):
            # 아래 문구가 있어야 list_team의 값이 다른 곳에 클릭을 해도 값이 유지가 됨
            list_person.get(list_person.curselection())
            
            df_person = pd.read_excel("C:/Python/Code/ToDoList/ToDoList_Form.xlsx",
                                    sheet_name = 'Person')
            div_person = df_person['분류'].drop_duplicates()
            div_person = div_person.reset_index()
            div_person = div_person.drop('index',axis=1)
            # for div in div_person["분류"]:
            #     list_person.insert(END,div)

            # list_team2 reset
            list_person2.delete(0,END)

            for row in range(len(df_person['분류'])):
                if df_person['분류'][row] == div_person['분류'][list_person.curselection()[0]]:
                    list_person2.insert(END,df_person['이름'][row])
        

        tree.bind('<ButtonRelease-1>', get_text)
        tree.bind('<ButtonRelease-1>', get_iid)

        # 상위 업무명/내용
        frame_upper_task_name = Frame(add_tk)
        frame_upper_task_name.pack(fill="x")

        lbl_upper_task = Label(frame_upper_task_name, text="상위 업무명/내용")
        lbl_upper_task.pack(side="left")
        
        lbl_upper_task_name = Label(frame_upper_task_name, text=get_text())
        lbl_upper_task_name.pack(side="right")
        
        # 업무명/내용
        frame_task_name = Frame(add_tk)
        frame_task_name.pack(fill="x")
        
        lbl_task_name = Label(frame_task_name, text="하위 추가 업무명/내용")
        lbl_task_name.pack(side="left")

        entry_task_name = Entry(frame_task_name)
        entry_task_name.pack(side="right")

        # 라디오버튼
        frame_radio = LabelFrame(add_tk, text="선택")
        frame_radio.pack(fill="x")

        # master를 추가 안 하면 무조건 0 값만 나옴 
        rad_var = IntVar(master=add_tk)
        rbt_one = Radiobutton(frame_radio, text="1) 단일 팀/담당자  ", value=1, variable=rad_var)
        rbt_teams = Radiobutton(frame_radio, text="2) 다수 팀  ", value=2, variable=rad_var)
        rbt_people = Radiobutton(frame_radio, text="3) 다수 담당자", value=3, variable=rad_var)

        # 기본값 선택
        rbt_one.select()
        # rbt_teams.select()

        rbt_one.pack(side="left")
        rbt_teams.pack(side="left")
        rbt_people.pack(side="left")


        # 담당팀(단수)
        frame_team_name = Frame(add_tk)
        frame_team_name.pack(fill="x")

        lbl_team_name = Label(frame_team_name, text="1) 담당팀(단일)")
        lbl_team_name.pack(side="left")

        entry_team_name = Entry(frame_team_name)
        entry_team_name.pack(side="right")

        # 담당팀(복수)
        frame_team_name2 = Frame(add_tk)
        # frame_team_name2.pack()
        frame_team_name2.pack(fill="x")

        lbl_team_name2 = Label(frame_team_name2, text="2) 담당팀(다수)")
        lbl_team_name2.pack(side="left")

        btn_team_edit = Button(frame_team_name2, text='즐겨찾기 편집(팀)', command=fav_team_edit)
        btn_team_edit.pack(side="right")

        # 담당팀(복수) - 리스트
        frame_team_name3 = Frame(add_tk)
        # frame_team_name3.pack()
        # frame_team_name3.pack(side="left")
        frame_team_name3.pack(fill="both")
        # frame_team_name3.pack(fill="x")

        yscrollbar = Scrollbar(frame_team_name3)
        yscrollbar.pack(side="left", fill='y')

        # xscrollbar x축의 절반만 나오게 하는 것 실패
        xscrollbar = Scrollbar(frame_team_name3, orient=HORIZONTAL)
        # xscrollbar.pack(side="bottom")
        # xscrollbar.pack(side="bottom",expand=True)
        # xscrollbar.pack(side="bottom", fill='x',expand=True)
        xscrollbar.pack(side="bottom", fill='x')
        
        # xscrollbar.place(relwidth=0.5)

        list_team = Listbox(frame_team_name3, selectmode="browse", height = 6,
        yscrollcommand=yscrollbar.set, xscrollcommand=xscrollbar.set)

        list_team.bind('<<ListboxSelect>>',get_team)        

        # list_team.pack(side="left", expand=1)
        # list_team.place(x=0, y=200)
        # list_team.pack(side="left", fill="y")
        list_team.pack(side="left", fill="both", expand=True)


        yscrollbar.config(command=list_team.yview)
        xscrollbar.config(command=list_team.xview)



        # 담당팀(복수) - 리스트
        # frame_team_name5 = Frame(add_tk)
        # frame_team_name5.pack()
        # frame_team_name5.pack(side="right")

        yscrollbar2 = Scrollbar(frame_team_name3)
        yscrollbar2.pack(side="right", fill='y')

        xscrollbar2 = Scrollbar(frame_team_name3, orient=HORIZONTAL)
        # xscrollbar2.pack(side="bottom")
        xscrollbar2.pack(side="bottom", fill='x',expand=True)

        list_team2 = Listbox(frame_team_name3, selectmode="extended", height = 6,
        yscrollcommand=yscrollbar2.set, xscrollcommand=xscrollbar2.set)

        # list_file = Listbox(list_frame, selectmode="extended", height = 10,
        # yscrollcommand=yscrollbar.set, xscrollcommand=xscrollbar.set, wrap=NONE)
        # list_team2.pack(side="left", expand=1)
        list_team2.pack(side="right", fill="both", expand=True)


        yscrollbar2.config(command=list_team2.yview)
        xscrollbar2.config(command=list_team2.xview)


        team_load()

        # 담당팀 복수 버튼
        # frame_team_name4 = Frame(add_tk)
        # frame_team_name4.pack(fill="x")
        # frame_team_name4.pack(fill="x")

        # btn_team_load = Button(frame_team_name4, text='즐겨찾기 불러오기', command = team_load)
        # btn_team_load.pack(side="left")


        # 담당자
        frame_person_name = Frame(add_tk)
        frame_person_name.pack(fill="x")

        lbl_person_name = Label(frame_person_name, text="1) 담당자")
        lbl_person_name.pack(side="left")

        entry_person_name = Entry(frame_person_name)
        entry_person_name.pack(side="right")

        # 담당자(다수)
        frame_person_name2 = Frame(add_tk)
        # frame_team_name2.pack()
        frame_person_name2.pack(fill="x")

        lbl_person_name2 = Label(frame_person_name2, text="3) 담당자(다수)")
        lbl_person_name2.pack(side="left")

        btn_person_edit = Button(frame_person_name2, text='즐겨찾기 편집(담당자)', command=fav_person_edit)
        btn_person_edit.pack(side="right")

        # 담당팀(복수) - 리스트
        frame_person_name3 = Frame(add_tk)
        frame_person_name3.pack(fill="both")
        # frame_team_name3.pack(fill="x")

        yscrollbar3 = Scrollbar(frame_person_name3)
        yscrollbar3.pack(side="left", fill='y')

        # xscrollbar x축의 절반만 나오게 하는 것 실패
        xscrollbar3 = Scrollbar(frame_person_name3, orient=HORIZONTAL)
        xscrollbar3.pack(side="bottom", fill='x')
        
        list_person = Listbox(frame_person_name3, selectmode="browse", height = 6,
        yscrollcommand=yscrollbar3.set, xscrollcommand=xscrollbar3.set)

        list_person.bind('<<ListboxSelect>>',get_person)        

        # list_team.pack(side="left", expand=1)
        # list_team.place(x=0, y=200)
        # list_team.pack(side="left", fill="y")
        list_person.pack(side="left", fill="both", expand=True)


        yscrollbar3.config(command=list_person.yview)
        xscrollbar3.config(command=list_person.xview)



        # 담당팀(복수) - 리스트

        yscrollbar4 = Scrollbar(frame_person_name3)
        yscrollbar4.pack(side="right", fill='y')

        xscrollbar4 = Scrollbar(frame_person_name3, orient=HORIZONTAL)
        # xscrollbar2.pack(side="bottom")
        xscrollbar4.pack(side="bottom", fill='x',expand=True)

        list_person2 = Listbox(frame_person_name3, selectmode="extended", height = 6,
        yscrollcommand=yscrollbar4.set, xscrollcommand=xscrollbar4.set)

        # list_file = Listbox(list_frame, selectmode="extended", height = 10,
        # yscrollcommand=yscrollbar.set, xscrollcommand=xscrollbar.set, wrap=NONE)
        # list_team2.pack(side="left", expand=1)
        list_person2.pack(side="right", fill="both", expand=True)


        yscrollbar4.config(command=list_person2.yview)
        xscrollbar4.config(command=list_person2.xview)


        person_load()


        # 상황
        frame_situation = Frame(add_tk)
        frame_situation.pack(fill="x")

        lbl_situation = Label(frame_situation, text="상황")
        lbl_situation.pack(side="left")

        opt_situation = ["진행","완료","검토","취소","중단","지연"]
        cmb_situation = ttk.Combobox(frame_situation,state="readonly",
        values=opt_situation)
        cmb_situation.current(0)
        cmb_situation.pack(side="right")

        # 완료날짜
        frame_date = Frame(add_tk)
        frame_date.pack(fill="x")

        lbl_date = Label(frame_date, text="완료날짜")
        lbl_date.pack(side="left")

        entry_date = Entry(frame_date)
        entry_date.pack(side="right")
        # entry_date.insert(END, "YY-MM-DD")

        # 완료시간
        frame_time = Frame(add_tk)
        frame_time.pack(fill="x")

        lbl_time = Label(frame_time, text="완료시간")
        lbl_time.pack(side="left")

        entry_time = Entry(frame_time)
        entry_time.pack(side="right")
        # entry_time.insert(END, "HH:MM")

        
        # 비고
        frame_note = Frame(add_tk)
        frame_note.pack(fill="x")

        lbl_note = Label(frame_note, text="비고")
        lbl_note.pack(side="left")

        entry_note = Entry(frame_note)
        entry_note.pack(side="right")



        # fucntion2 frame
        frame_func2 = Frame(add_tk)
        frame_func2.pack(fill="x")

        # add db
        btn_add_db = Button(frame_func2, text="추가",command=add_db)
        btn_add_db.pack(side="left")

        # cancel button
        # quit 함수를 쓰면 전체 프로그램이 종료되어서 withdraw 함수 사용함
        btn_cancel = Button(frame_func2, text="취소", command=add_tk.withdraw)
        btn_cancel.pack(side="right")

        add_tk.mainloop()

    except Exception as err:
        msgbox.showerror("에러", err)


def del_task():
    try:
        # print(get_iid())
        wb = load_workbook("C:/Python/Code/ToDoList/ToDoList_Form.xlsx")
        db_ws = wb['DB']
        # values_only=True를 해야 위치값이 아닌 실제 데이터 값이 도출됨
        db_header = db_ws.iter_rows(min_row=1, max_row=1, max_col=9, values_only=True)
        db_datas = db_ws.iter_rows(min_row=2, max_col=9, values_only=True)
        db_header = [r for r in db_header]
        db_datas = [r for r in db_datas]
        
        # 선택 항목의 iid값 가져오기
        for i in range(1, db_ws.max_row+1):
            if db_ws.cell(i,1).value == get_iid():
                # print(i)
                r_selected = i
        
        # db_ws.delete_rows(r_selected)

        # 선택 값이 1레벨이면 하위 2,3레벨도 삭제
        if db_datas[r_selected-2][0][3:8] == "00-00":
            for x in reversed(range(db_ws.max_row-1)):
                if db_datas[x][0][0:2] == db_datas[r_selected-2][0][0:2]:
                    db_ws.delete_rows(x+2)
        # 선택 값이 2레벨이면 3레벨도 삭제
        elif db_datas[r_selected-2][0][6:8] == "00":
            for x in reversed(range(db_ws.max_row-1)):
                if db_datas[x][0][0:5] == db_datas[r_selected-2][0][0:5]:
                    db_ws.delete_rows(x+2)
        # 선택 값이 3레벨이면 해당 행만 삭제
        else:
            db_ws.delete_rows(r_selected)

        wb.save("C:/Python/Code/ToDoList/ToDoList_Form.xlsx")
        
        msgbox.showinfo("알림", "선택한 업무 삭제 되었습니다.")
        sort()
        load_task()

    except Exception as err:
        msgbox.showerror("에러", err)

def edit_task():
    try:
        add_tk = Tk()
        add_tk.title("수정")


        def apply():
            # print(entry_task_name.get())
                
            wb = load_workbook("C:/Python/Code/ToDoList/ToDoList_Form.xlsx")
            db_ws = wb['DB']    
            # values_only=True를 해야 위치값이 아닌 실제 데이터 값이 도출됨
            # db_header = db_ws.iter_rows(min_row=1, max_row=1, max_col=9, values_only=True)
            db_datas = db_ws.iter_rows(min_row=2, max_col=9, values_only=True)
            # db_header = [r for r in db_header]
            db_datas = [r for r in db_datas]
            
            
            # 선택 항목의 iid값 가져오기
            for i in range(1, db_ws.max_row+1):
                if db_ws.cell(i,1).value == get_iid():
                    # print(i)
                    # r_selected = i

                    if len(entry_task_name.get()) == 0:
                        db_ws.cell(row=i, column=2).value = " "
                    else :    
                        db_ws.cell(row=i, column=2).value = entry_task_name.get()
                    if len(entry_team_name.get()) == 0 :
                        db_ws.cell(row=i, column=3).value = " "
                    else :    
                        db_ws.cell(row=i, column=3).value = entry_team_name.get()
                    if len(entry_person_name.get()) == 0:
                        db_ws.cell(row=i, column=4).value = " "
                    else :    
                        db_ws.cell(row=i, column=4).value = entry_person_name.get()
                    db_ws.cell(row=i, column=5).value = cmb_situation.get()
                    if len(entry_date.get()) == 0:
                        db_ws.cell(row=i, column=6).value = " "    
                    else :
                        db_ws.cell(row=i, column=6).value = entry_date.get()
                    if len(entry_time.get()) == 0:
                        db_ws.cell(row=i, column=7).value = " "
                    else :    
                        db_ws.cell(row=i, column=7).value = entry_time.get()
                    if len(entry_note.get()) == 0:
                        db_ws.cell(row=i, column=8).value = " "
                    else :     
                        db_ws.cell(row=i, column=8).value = entry_note.get()

            wb.save("C:/Python/Code/ToDoList/ToDoList_Form.xlsx")

            msgbox.showinfo("알림", "수정 되었습니다.")

            add_tk.withdraw()
            sort()
            load_task()

        # print(get_iid())

        wb = load_workbook("C:/Python/Code/ToDoList/ToDoList_Form.xlsx")
        db_ws = wb['DB']    
        # values_only=True를 해야 위치값이 아닌 실제 데이터 값이 도출됨
        # db_header = db_ws.iter_rows(min_row=1, max_row=1, max_col=9, values_only=True)
        db_datas = db_ws.iter_rows(min_row=2, max_col=9, values_only=True)
        # db_header = [r for r in db_header]
        db_datas = [r for r in db_datas]
        
        
        # 선택 항목의 iid값 가져오기
        for i in range(1, db_ws.max_row+1):
            if db_ws.cell(i,1).value == get_iid():
                # print(i)
                r_selected = i
        # print(db_datas[r_selected-2][1])
        
        task_name = db_datas[r_selected-2][1]
        # print(task_name)
        team_name = db_datas[r_selected-2][2]
        person_name = db_datas[r_selected-2][3]
        situation = db_datas[r_selected-2][4]
        d_date = db_datas[r_selected-2][5]
        d_time = db_datas[r_selected-2][6]
        note = db_datas[r_selected-2][7]


        # for x in range(db_ws.max_row-1):

        #     if db_datas[x][0] == db_datas[r_selected-2][0]:
                # print(db_datas[x][0])
                # print(db_datas[r_selected-2][0])

        # 업무명/내용
        frame_task_name = Frame(add_tk)
        frame_task_name.pack(fill="x")
        
        lbl_task_name = Label(frame_task_name, text="업무명/내용")
        lbl_task_name.pack(side="left")
        
        entry_task_name = Entry(frame_task_name)
        # 기존 DB값 불러오기
        entry_task_name.insert(0,task_name)

        entry_task_name.pack(side="right")

        # 담당팀
        frame_team_name = Frame(add_tk)
        frame_team_name.pack(fill="x")

        lbl_team_name = Label(frame_team_name, text="담당팀")
        lbl_team_name.pack(side="left")

        entry_team_name = Entry(frame_team_name)
        entry_team_name.insert(0,team_name)
        entry_team_name.pack(side="right")

        # 담당자
        frame_person_name = Frame(add_tk)
        frame_person_name.pack(fill="x")

        lbl_person_name = Label(frame_person_name, text="담당자")
        lbl_person_name.pack(side="left")

        entry_person_name = Entry(frame_person_name)
        entry_person_name.insert(0,person_name)
        entry_person_name.pack(side="right")

        # 상황
        frame_situation = Frame(add_tk)
        frame_situation.pack(fill="x")

        lbl_situation = Label(frame_situation, text="상황")
        lbl_situation.pack(side="left")

        opt_situation = ["진행","완료","검토","취소","중단","지연"]
        cmb_situation = ttk.Combobox(frame_situation,state="readonly",
        values=opt_situation)
        cmb_situation.current(0)
        cmb_situation.pack(side="right")

        # 완료날짜
        frame_date = Frame(add_tk)
        frame_date.pack(fill="x")

        lbl_date = Label(frame_date, text="완료날짜")
        lbl_date.pack(side="left")

        entry_date = Entry(frame_date)
        entry_date.insert(0,d_date)
        entry_date.pack(side="right")
        # entry_date.insert(END, "YY-MM-DD")

        # 완료시간
        frame_time = Frame(add_tk)
        frame_time.pack(fill="x")

        lbl_time = Label(frame_time, text="완료시간")
        lbl_time.pack(side="left")

        entry_time = Entry(frame_time)
        entry_time.insert(0,d_time)
        entry_time.pack(side="right")
        # entry_time.insert(END, "HH:MM")

        
        # 비고
        frame_note = Frame(add_tk)
        frame_note.pack(fill="x")

        lbl_note = Label(frame_note, text="비고")
        lbl_note.pack(side="left")

        entry_note = Entry(frame_note)
        entry_note.insert(0,note)
        entry_note.pack(side="right")


        # fucntion2 frame
        frame_func2 = Frame(add_tk)
        frame_func2.pack(fill="x")

        # add db
        btn_apply = Button(frame_func2, text="수정반영",command=apply)
        btn_apply.pack(side="left")

        # cancel button
        # quit 함수를 쓰면 전체 프로그램이 종료되어서 withdraw 함수 사용함
        btn_cancel = Button(frame_func2, text="취소", command=add_tk.withdraw)
        btn_cancel.pack(side="right")

        add_tk.mainloop()

    except Exception as err:
        msgbox.showerror("에러", err)

def load_task():
    try:
        wb = load_workbook("C:/Python/Code/ToDoList/ToDoList_Form.xlsx")
        db_ws = wb['DB']    
        # values_only=True를 해야 위치값이 아닌 실제 데이터 값이 도출됨
        db_header = db_ws.iter_rows(min_row=1, max_row=1, max_col=9, values_only=True)
        db_datas = db_ws.iter_rows(min_row=2, max_col=9, values_only=True)
        db_header = [r for r in db_header]
        db_datas = [r for r in db_datas]
        
        # treeview reset
        # 이게 없으면 업무 추가시 기존 데이터가 2번 돌면서 iid가 또 있다고 하면서 에러가 남
        tree.delete(*tree.get_children())

        for data in db_datas:
            # print(data[0][6:8])
            # text 값이 있어야 하이라키구조 처럼 보여질 수 있음 (+-표시).
            # lv1에만 open=true를 줘서 실행 시 lv2까지만 자동 보여주기    
            if data[0][3:8] == '00-00':
                lv1 = tree.insert('','end', iid=data[0], text=data[1], values=data[2:8])
                # lv1 = tree.insert('','end', iid=data[0], text=data[1], values=data[2:8], open=True)
            elif '00' in data[0][6:8]:
                lv2 = tree.insert(lv1,'end',iid=data[0], text=data[1] ,values=data[2:8])
            else :
                lv3 = tree.insert(lv2,'end',iid=data[0], text=data[1] ,values=data[2:8])

    except Exception as err:
        msgbox.showerror("에러", err)



def sort():
    df_db = pd.read_excel("C:/Python/Code/ToDoList/ToDoList_Form.xlsx", sheet_name='DB')
    # print(db_df)
    df_db = df_db.sort_values(by=['No'])

    # 넘버링 새롭게 하기
    # 레벨2, 3는 아직 못함. 레벨1만 넘버링함
    lv1_cnt = 0
    for i in range(len(df_db['업무레벨'])):
        # print(df_db['업무레벨'][i])
        if df_db['업무레벨'][i] == 1 :
            lv1_cnt += 1
    # print(lv1_cnt)
    # print(int(df_db['No'][len(df_db['No'])-1][0:2]))
    # print(type(lv1_cnt))
    # if lv1_cnt == 7:
    #     print("123")
    # if int(df_db['No'][len(df_db['No'])-1][0:2]) == 7:
    #     print("445")
    lv1_cnt = 0
    if lv1_cnt != int(df_db['No'][len(df_db['No'])-1][0:2]) :
        # print("different")
        for i in range(len(df_db['업무레벨'])):
        # print(df_db['업무레벨'][i])
            if df_db['업무레벨'][i] == 1 :
                lv1_cnt += 1
            # print("asis : ", df_db['No'][i])
            df_db['No'][i] = format(lv1_cnt,'02') + "-"+ df_db['No'][i][3:]
            # print("tobe : ", df_db['No'][i])
            
            # df_db['No'][len(df_db['No'])-1][0:2] = format(lv1_cnt,'02')
                
    # print("lv1_cnt : ", lv1_cnt)
    # print(db_df)
    df_db.to_excel("C:/Python/Code/ToDoList/test.xlsx",index=False, sheet_name="DB")
    path1 = "C:/Python/Code/ToDoList/ToDoList_Form.xlsx"
    path2 = "C:/Python/Code/ToDoList/test.xlsx"
    
    wb2 = load_workbook(filename=path2)
    ws2 = wb2['DB']

    wb1 = load_workbook(filename=path1)
    wb1.remove(wb1['DB'])
    ws1 = wb1.create_sheet()
    ws1.title = 'DB'
    # ws1 = wb1['DB']
    for row in ws2:
        for cell in row:
            ws1[cell.coordinate].value = cell.value
    wb1.save(path1)

win = Tk()
win.geometry("800x500")
win.title("To Do List")

# font
func_font = font.Font(family='맑은 고딕', size=12)
func_font2 = font.Font(family='맑은 고딕', size=10)
# func_font = font.Font(family='맑은 고딕', size=12, weight='bold')


# browse : 1개만 선택
# extended : 2개 이상 선택 가능
# tree = ttk.Treeview(win, selectmode='extended')
tree = ttk.Treeview(win, selectmode='browse')
tree.pack(expand=True, fill="both")


style = ttk.Style()
# style.configure("Treview.Heading", font=('맑은 고딕',100))
style.configure(".", font=('맑은 고딕',12), rowheight=30)

# # Number of rows to display, default is 10
tree['height'] = 20

tree['columns'] = ("1", "2", "3", "4", "5", "6")

# column보다 heading이 1개 더 많아서(#0) 하이라키구조 처럼 +- 표시가 됨. 
tree.heading("#0", text="업무명/내용")
tree.heading("#1", text="담당팀")
tree.heading("#2", text="담당자")
tree.heading("#3", text="상황")
tree.heading("#4", text="완료날짜")
tree.heading("#5", text="완료시간")
tree.heading("#6", text="비고")

tree.column("1", width=100)
tree.column("2", width=100)
tree.column("3", width=50, anchor='c')
tree.column("4", width=70, anchor='c')
tree.column("5", width=70, anchor='c')
tree.column("6", width=100)



yscrollbar = Scrollbar(tree, orient="vertical", command=tree.yview)
yscrollbar.pack(side="right", fill='y')
tree.config(yscrollcommand=yscrollbar.set)


# x스크롤바 실행 실패함.
# xscrollbar = Scrollbar(tree, orient='horizontal', command=tree.xview)
# xscrollbar = Scrollbar(tree, orient='horizontal')

# # xscrollbar = Scrollbar(tree, orient=HORIZONTAL, command=tree.xview)
# xscrollbar.pack(side="bottom", fill='x')

# tree.config(xscrollcommand=xscrollbar.set)
# xscrollbar.configure(command=tree.xview)

# tree.config(xscrollcommand=xscrollbar.set)


# yscrollbar.config(command=tree.yview)
# xscrollbar.config(command=tree.xview)


# function frame
# func_frame = Frame(tree)
func_frame = Frame(win)

func_frame.pack(side="bottom", fill="x", padx=5, pady=5)
# func_frame.pack(fill="x")

# 업무 추가 버튼
btn_add_task = Button(func_frame, text="신규 추가(Lv1)", padx=5, pady=5, command=add_task)
btn_add_task['font'] = func_font
btn_add_task.pack(side="left", padx=5)

# # 하위 업무 추가
btn_add_task2 = Button(func_frame, text="하위 추가(Lv2~3)", command=add_task2, padx=5, pady=5)
btn_add_task2['font'] = func_font
btn_add_task2.pack(side="left", padx=5)

# 삭제 버튼
btn_del_task = Button(func_frame, text="삭제", command=del_task, padx=5, pady=5)
btn_del_task['font'] = func_font
btn_del_task.pack(side="left", padx=5)

# 편집 버튼
btn_edit_task = Button(func_frame, text="수정", command=edit_task, padx=5, pady=5)
btn_edit_task['font'] = func_font
btn_edit_task.pack(side="left", padx=5)

# 종료 버튼
btn_close = Button(func_frame, text="종료", command=win.quit, padx=5, pady=5)
btn_close['font'] = func_font
btn_close.pack(side="right", padx=5)

sort()
load_task()

win.mainloop()

