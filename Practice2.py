from openpyxl import load_workbook
wb = load_workbook(filename="C:/Python/Code/ToDoList/student.xlsx", read_only=False)
ws = wb['student'] # connecting to work sheet
# l1=ws.iter_rows(min_row=1,max_row=1,max_col=5)
l1=ws.iter_rows(min_row=1,max_row=1,max_col=5,values_only=True)

r_set=ws.iter_rows(min_row=2,max_row=5,values_only=True)
#print(list(l1))
l1=[r for r in l1] # Prepare list for column headers 
r_set=[r for r in r_set] # Prepare list with data 
wb.close()# Close the workbook after reading
#print(l1) # to check the headers 

from tkinter import ttk # for treeview 
import tkinter as tk
my_w = tk.Tk() # Main window 
my_w.geometry("560x280") # width and hight of window 
my_w.title("www.plus2net.com")  
# Using treeview widget
trv = ttk.Treeview(my_w, selectmode ='browse')
trv.grid(row=0,column=0,columnspan=3,padx=30,pady=20)

trv['height']=5 # Number of rows to display, default is 10
trv['show'] = 'headings' 
# column identifiers 
trv["columns"] = l1[0]
# Defining headings, other option in tree
# width of columns and alignment 
for i in l1[0]:
    trv.column(i, width = 100, anchor ='c')
# Headings of respective columns
for i in l1[0]:
    trv.heading(i, text =i)

## Adding data to treeview 
for dt in r_set:  
    trv.insert("",'end',iid=dt[0],values=dt) # adding row
my_w.mainloop()